"""Export the Bangkok district site-evaluation into a formatted .xlsx workbook.

Merges every per-district criterion we have into one review-ready spreadsheet:
  - Tier 1 MCDA result (ranked, weighted score + normalized sub-scores)
  - the raw criteria behind each score
  - the Tier 0 work<->home intercept metrics
  - full seasonal UHII detail (3 seasons x night/evening)

Sources (all keyed on District, 50 rows each):
  data/gis/site_selection_scores.csv     (build_site_selection_matrix.py)
  data/gis/bangkok_uhi_data.csv          (merge_uhi_data.py)
  data/gis/bangkok_transit_data.csv
  data/gis/bangkok_intercept_scores.csv  (fetch_land_use_osm.py)

Output: data/gis/site_district_evaluation.xlsx
"""
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference

# MCDA weights (must match build_site_selection_matrix.py)
WEIGHTS = {"uhi": 0.35, "vuln": 0.30, "transit": 0.25, "pop": 0.10}

GIS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gis")
OUT = os.path.join(GIS, "site_district_evaluation.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
GROUP_FILL = PatternFill("solid", fgColor="2E75B6")
TOP3_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_merged():
    s = pd.read_csv(os.path.join(GIS, "site_selection_scores.csv"))
    u = pd.read_csv(os.path.join(GIS, "bangkok_uhi_data.csv"))
    t = pd.read_csv(os.path.join(GIS, "bangkok_transit_data.csv"))
    i = pd.read_csv(os.path.join(GIS, "bangkok_intercept_scores.csv"))
    return (s.merge(u, on="District", how="left")
             .merge(t, on="District", how="left")
             .merge(i, on="District", how="left"))


def main():
    m = load_merged().sort_values("Rank")

    # ---- Sheet 1: Site Evaluation (ranked, all criteria) ------------------
    main_cols = [
        ("Rank", "Rank"),
        ("District", "District"),
        ("UHI_Tier", "UHI Tier"),
        ("score", "TOTAL SCORE (0-10)"),
        ("s_uhi", "UHI score (35%)"),
        ("s_vuln", "Vulnerability score (30%)"),
        ("s_transit", "Transit score (25%)"),
        ("s_pop", "Population score (10%)"),
        ("UHII_HotDry_Evening_C", "Evening UHII HotDry (C)"),
        ("LST_Mean_C", "LST mean (C)"),
        ("LST_Max_C", "LST max (C)"),
        ("Dominant_LCZ", "Dominant LCZ"),
        ("LCZ_Confidence_Pct", "LCZ conf (%)"),
        ("Transit_Station_Count", "Transit stations"),
        ("Population_Pct_BMA", "Population (% BMA)"),
        ("Intercept_Score_Pct", "Tier0 Intercept (% district)"),
        ("Intercept_Pct_of_Residential", "Tier0 Intercept (% resid.)"),
    ]
    main_df = m[[c for c, _ in main_cols]].copy()
    main_df.columns = [h for _, h in main_cols]

    # ---- Sheet 2: Seasonal UHII detail (+ diagnostic mean/range) ----------
    ev = ["UHII_CoolDry_Evening_C", "UHII_HotDry_Evening_C", "UHII_Wet_Evening_C"]
    seas = m[["District"] + ev +
             ["UHII_CoolDry_Night_C", "UHII_HotDry_Night_C", "UHII_Wet_Night_C"]].copy()
    seas["Eve_Mean (diagnostic)"] = m[ev].mean(axis=1).round(2)
    seas["Eve_Range (diagnostic)"] = (m[ev].max(axis=1) - m[ev].min(axis=1)).round(2)
    seas.columns = ["District", "CoolDry Eve", "HotDry Eve", "Wet Eve",
                    "CoolDry Night", "HotDry Night", "Wet Night",
                    "Eve Mean (diag)", "Eve Range (diag)"]

    # ---- Sheet 3: Score breakdown (weighted points toward total) ----------
    # contribution_i = weight_i x normalized_sub_score_i ; the four sum to TOTAL.
    brk = pd.DataFrame({
        "Rank": m["Rank"],
        "District": m["District"],
        "UHI pts (35%)": (WEIGHTS["uhi"] * m["s_uhi"]).round(2),
        "Vulnerability pts (30%)": (WEIGHTS["vuln"] * m["s_vuln"]).round(2),
        "Transit pts (25%)": (WEIGHTS["transit"] * m["s_transit"]).round(2),
        "Population pts (10%)": (WEIGHTS["pop"] * m["s_pop"]).round(2),
        "TOTAL": m["score"].round(2),
    }).sort_values("Rank").reset_index(drop=True)

    # ---- Sheet 4: Legend & method ----------------------------------------
    legend = [
        ["Bangkok District Site Evaluation - Legend & Method", ""],
        ["", ""],
        ["Purpose", "Tier 1 weighted multi-criteria ranking of Bangkok's 50 districts "
                    "for an evening thermal-refuge civic center."],
        ["", ""],
        ["MCDA criteria & weights", ""],
        ["  UHI severity (evening)", "35%  = 0.7 x Evening UHII (HotDry) + 0.3 x LST mean, min-max normalized 0-10"],
        ["  Housing vulnerability", "30%  = Dominant LCZ ordinal (LCZ3=10, LCZ2=6, LCZ8=4, else 2)"],
        ["  Transit / intercept", "25%  = rapid-transit station count, min-max normalized 0-10"],
        ["  Demographic density", "10%  = Population % of BMA, min-max normalized 0-10"],
        ["  TOTAL SCORE", "weighted sum of the four sub-scores, 0-10"],
        ["", ""],
        ["Tier 0 (spatial gate)", "Work<->home intercept: residential fabric within 200 m of a working "
                                  "zone. Currently reported alongside, NOT yet folded into the score."],
        ["  Intercept (% district)", "share of the district area that is intercept fabric"],
        ["  Intercept (% resid.)", "share of the district's residential fabric that is intercept-adjacent"],
        ["", ""],
        ["Seasonal UHII sheet", "Full 3-season x night/evening urban heat-island intensity (C)."],
        ["  Eve Mean/Range (diag)", "DIAGNOSTIC ONLY - not part of the score. Range = seasonal "
                                    "consistency (low = hot every evening season). See "
                                    "docs/plan_evening-uhi-seasonal-average.md."],
        ["", ""],
        ["Caveats", "LST is a ~10:30 daytime SURFACE signal (heat-storage proxy, not evening air temp). "
                    "UHI_Tier is derived from night UHII. OSM intercept = incomplete proxy. "
                    "Full method: docs/SiteSelectionMatrixGemini.md."],
        ["", ""],
        ["Generated by", "data/export_site_evaluation_xlsx.py"],
    ]
    legend_df = pd.DataFrame(legend, columns=["Field", "Definition"])

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        main_df.to_excel(writer, sheet_name="Site Evaluation", index=False)
        seas.to_excel(writer, sheet_name="Seasonal UHII", index=False)
        brk.to_excel(writer, sheet_name="Score Breakdown", index=False)
        legend_df.to_excel(writer, sheet_name="Legend & Method", index=False)

        wb = writer.book
        _format_table(wb["Site Evaluation"], main_df, top3=True, score_cols=range(4, 8))
        _format_table(wb["Seasonal UHII"], seas, top3=False, score_cols=None)
        _format_table(wb["Score Breakdown"], brk, top3=True, score_cols=[7])
        _add_breakdown_chart(wb["Score Breakdown"], brk, top_n=10)
        _format_legend(wb["Legend & Method"])

    print("Wrote", OUT)
    print("Top 3:", ", ".join(main_df["District"].head(3)))


def _format_table(ws, df, top3, score_cols):
    ncol = df.shape[1]
    # header
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "C2"  # lock header row + Rank/District (or District col)

    # column widths + borders + number formats
    for c in range(1, ncol + 1):
        col = get_column_letter(c)
        header = str(df.columns[c - 1])
        width = max(10, min(30, len(header) + 2))
        if header == "District":
            width = 24
        elif "LCZ" in header and "Dominant" in header:
            width = 22
        ws.column_dimensions[col].width = width
        for r in range(2, df.shape[0] + 2):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if isinstance(cell.value, float):
                cell.number_format = "0.00"
            if c > 2:
                cell.alignment = Alignment(horizontal="center")

    # highlight top-3 rows
    if top3:
        for r in range(2, 5):
            for c in range(1, ncol + 1):
                ws.cell(row=r, column=c).fill = TOP3_FILL

    # color scale on the total score + sub-scores
    if score_cols is not None:
        last = df.shape[0] + 1
        # total score col is index 4 (1-based) -> column D
        for cidx in [4] + list(score_cols):
            col = get_column_letter(cidx)
            ws.conditional_formatting.add(
                f"{col}2:{col}{last}",
                ColorScaleRule(start_type="min", start_color="F8696B",
                               mid_type="percentile", mid_value=50, mid_color="FFEB84",
                               end_type="max", end_color="63BE7B"))


def _add_breakdown_chart(ws, df, top_n=10):
    """Stacked horizontal bar of the four weighted contributions for the top N
    districts -- bar length == total score, segments == each criterion's points."""
    top_n = min(top_n, df.shape[0])
    chart = BarChart()
    chart.type = "bar"            # horizontal
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Top {} districts - weighted score contribution".format(top_n)
    chart.y_axis.title = "Points toward total (0-10)"
    chart.x_axis.title = "District"
    chart.height = 9
    chart.width = 22

    # The four contribution columns are C..F (cols 3..6); series titles from row 1.
    data = Reference(ws, min_col=3, max_col=6, min_row=1, max_row=top_n + 1)
    cats = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=top_n + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend.position = "b"
    ws.add_chart(chart, "I2")


def _format_legend(ws):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F4E78")
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1)
        b = ws.cell(row=r, column=2)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        if a.value and not a.value.startswith("  ") and r > 1:
            a.font = Font(bold=True)
    ws.sheet_view.showGridLines = False


if __name__ == "__main__":
    main()
